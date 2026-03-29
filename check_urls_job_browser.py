import os
import time
import sys

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("Please install openpyxl")
    sys.exit(1)

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
except ImportError:
    print("Please install playwright: pip install playwright && playwright install")
    sys.exit(1)

import requests

ENV_FILE = r"d:\Subscribtions Sellers\.env"
SUPABASE_URL = ""
SUPABASE_KEY = ""

try:
    with open(ENV_FILE, "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("SUPABASE_URL="):
                SUPABASE_URL = line.strip().split("=", 1)[1]
            elif line.startswith("SUPABASE_KEY="):
                SUPABASE_KEY = line.strip().split("=", 1)[1]
except Exception as e:
    print(f"Failed to read .env file: {e}")
    sys.exit(1)

EXCEL_FILE = r"D:\Ads_Links_Status.xlsx"

def append_to_excel(name_ar, name_en, plan_name, url, status, checked_mark):
    try:
        # Append without over-writing existing check
        if not os.path.exists(EXCEL_FILE):
             wb = Workbook()
             ws = wb.active
             ws.title = "Ads Links"
             ws.append(["Name (Arabic)", "Name (English)", "Plan / Duration", "URL", "Status", "Checked"])
        else:
             wb = load_workbook(EXCEL_FILE)
             ws = wb.active
             
        ws.append([name_ar, name_en, plan_name, url, status, checked_mark])
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Error saving to Excel for {name_en}: {e}")

def check_url_with_browser(page, url):
    if not url or str(url).strip() == '':
        return "No URL", "No"
    
    try:
        # Navigate using Playwright
        resp = page.goto(url, timeout=30000, wait_until="domcontentloaded")
        
        # Wait for Cloudflare bypass if present
        try:
            page.wait_for_timeout(3000)
            title = page.title()
            if "Just a moment..." in title or "Cloudflare" in title:
                page.wait_for_timeout(5000) # extra wait
        except:
            pass
            
        title = page.title().lower()
        active_url = page.url.lower()
        
        # Dead products typically redirect to 404
        if "z2u.com/404" in active_url or "not found" in title or "404" in title or "page not found" in title:
            return "Dead / Not Found (404)", "✔️ Yes (Browser)"
            
        # Maybe Z2U title contains "Z2U.com" or the product name
        if resp and resp.status == 200:
            return f"Working", "✔️ Yes (Browser)"
            
        status = resp.status if resp else "Unknown"
        if status == 403:
             # Even with browser it gave 403?
             if "z2u.com" in title or "just a moment" in title:
                 return f"Working / Pending CF ({status})", "✔️ Yes (Browser)"
                 
        return f"Working (Status: {status} / {title[:15]})", "✔️ Yes (Browser)"
        
    except PlaywrightTimeoutError:
        return "Timeout (Took too long)", "✔️ Yes (Browser)"
    except Exception as e:
        err = str(e).split('\n')[0][:30]
        return f"Failed ({err})", "✔️ Yes (Browser)"

def main():
    print("Starting URL check process with Playwright Browser...")
        
    url = f"{SUPABASE_URL}/rest/v1/products?select=*"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    
    try:
        resp = requests.get(url, headers=headers)
        products = resp.json()
    except:
        products = []

    if products:
        print(f"Fetched {len(products)} products from Supabase.")
    else:
        print("No products fetched. Exiting.")
        return

    processed_urls_per_product = {}

    with sync_playwright() as p:
        # Launch Chromium. Headless=False because the user wants "like a human"
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = context.new_page()

        for prod in products:
            p_id = prod.get('id', 'unknown')
            name_ar = prod.get('name_ar', '')
            name_en = prod.get('name_en', '')
            
            print(f"\nProcessing Product: {name_ar} ({name_en})")
            processed_urls_per_product[p_id] = set()
            
            main_url = prod.get('source_url')
            if main_url:
                print(f" -> Browser checking Main URL...")
                status, chk = check_url_with_browser(page, main_url)
                append_to_excel(name_ar, name_en, "Main Link", main_url, status, chk)
                print(f"    Result: {status}")
                processed_urls_per_product[p_id].add(main_url)
                
            plans = prod.get('subscription_plans', [])
            if plans and isinstance(plans, list):
                for plan in plans:
                    plan_ar = plan.get('label_ar', '')
                    plan_en = plan.get('label_en', '')
                    plan_url = plan.get('source_url')
                    plan_name = f"{plan_ar} / {plan_en}"
                    
                    if plan_url:
                        if plan_url in processed_urls_per_product[p_id]:
                            print(f" -> Skipping already checked URL.")
                            append_to_excel(name_ar, name_en, plan_name, plan_url, "Already Checked", "✔️ Yes (Browser)")
                        else:
                            print(f" -> Browser checking Plan URL: {plan_name}...")
                            status, chk = check_url_with_browser(page, plan_url)
                            append_to_excel(name_ar, name_en, plan_name, plan_url, status, chk)
                            processed_urls_per_product[p_id].add(plan_url)
                            print(f"    Result: {status}")
                            
        browser.close()
    print("\n✅ All products and URLs have been checked using the browser and saved.")
    print("Done! You can check D:\\Ads_Links_Status.xlsx")

if __name__ == '__main__':
    main()
