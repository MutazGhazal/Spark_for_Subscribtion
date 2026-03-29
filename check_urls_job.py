import os
import time
import requests
import json
import sys

# Try importing openpyxl, if missing, print error
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("openpyxl is not installed. Please install it using: pip install openpyxl requests")
    sys.exit(1)

ENV_FILE = r"d:\Subscribtions Sellers\.env"
SUPABASE_URL = ""
SUPABASE_KEY = ""

try:
    with open(ENV_FILE, "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("SUPABASE_URL="):
                SUPABASE_URL = line.strip().split("=", 1)[1]
            elif line.startswith("SUPABASE_KEY="):
                SUPABASE_KEY = line.strip().split("=", 1)[1]
except Exception as e:
    print(f"Failed to read .env file: {e}")
    sys.exit(1)

EXCEL_FILE = r"D:\Ads_Links_Status.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Ads Links"
        ws.append(["Name (Arabic)", "Name (English)", "Plan / Duration", "URL", "Status", "Checked"])
        wb.save(EXCEL_FILE)

def append_to_excel(name_ar, name_en, plan_name, url, status, checked_mark):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([name_ar, name_en, plan_name, url, status, checked_mark])
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Error saving to Excel for {name_en}: {e}")

def check_url(url):
    if not url or str(url).strip() == '':
        return "No URL", "No"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    try:
        r = requests.get(url, timeout=15, headers=headers)
        if r.status_code == 200:
            return "Working (200)", "✔️ Yes"
        elif r.status_code == 403:
            return "Working / bot-protection (403)", "✔️ Yes"
        elif r.status_code == 404:
            return "Not Found (404)", "✔️ Yes"
        else:
            return f"Status: {r.status_code}", "✔️ Yes"
    except Exception as e:
        err = str(e).split(':')[-1].strip()[:30]
        return f"Failed ({err})", "✔️ Yes"

def main():
    print("Starting URL check process...")
    try:
        init_excel()
        print(f"Excel initialized at {EXCEL_FILE}")
    except Exception as e:
        print("Failed to initialize Excel:", e)
        return
        
    print("Fetching products from Supabase...")
    url = f"{SUPABASE_URL}/rest/v1/products?select=*"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json"
    }
    
    try:
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        products = resp.json()
    except Exception as e:
        print("Failed to fetch products from Supabase:", e)
        # Fallback empty list to avoid crash
        products = []

    if products:
        print(f"Fetched {len(products)} products from Supabase.")
    else:
        print("No products fetched. Exiting.")
        return

    processed_urls_per_product = {}

    for p in products:
        p_id = p.get('id', 'unknown')
        name_ar = p.get('name_ar', '')
        name_en = p.get('name_en', '')
        
        print(f"\nProcessing Product: {name_ar} ({name_en})")
        processed_urls_per_product[p_id] = set()
        
        main_url = p.get('source_url')
        if main_url:
            print(f" -> Checking Main URL...")
            status, chk = check_url(main_url)
            append_to_excel(name_ar, name_en, "Main Link", main_url, status, chk)
            print(f"    Result: {status}")
            processed_urls_per_product[p_id].add(main_url)
            time.sleep(1)
            
        plans = p.get('subscription_plans', [])
        if plans and isinstance(plans, list):
            for plan in plans:
                plan_ar = plan.get('label_ar', '')
                plan_en = plan.get('label_en', '')
                plan_url = plan.get('source_url')
                plan_name = f"{plan_ar} / {plan_en}"
                
                if plan_url:
                    if plan_url in processed_urls_per_product[p_id]:
                        print(f" -> Skipping already checked URL for {plan_name}...")
                        append_to_excel(name_ar, name_en, plan_name, plan_url, "Already Checked above", "✔️ Yes")
                    else:
                        print(f" -> Checking Plan URL: {plan_name}...")
                        status, chk = check_url(plan_url)
                        append_to_excel(name_ar, name_en, plan_name, plan_url, status, chk)
                        processed_urls_per_product[p_id].add(plan_url)
                        print(f"    Result: {status}")
                        time.sleep(1)
                        
    print("\n✅ All products and URLs have been checked and saved to Excel.")
    print("Excel File is located at:", EXCEL_FILE)

if __name__ == '__main__':
    main()
